import openpyxl
from openpyxl import styles
from openpyxl.styles.borders import Border, Side

# path = 'sample.xlsx'
#
# workbook = openpyxl.Workbook()
#
# # workbook.save(filename=path)
#
# sheet = workbook.active

#
# bold_style = styles.NamedStyle(name='header')
# bold_style.font = styles.Font(bold=True, size=11)
# bold_style.alignment = styles.Alignment(horizontal='center', vertical='center')
#
#
# sheet.merge_cells(f'A{sheet.max_row}:D{sheet.max_row+1}')
#
# data = [
#     [
#         ('',),
#         ('№', 'FIO', 'Summa', '40%',  '10%'),
#         (1, 'Abduraxmatov Mirzoxid', 100000, 40000,  10000),
#         (2, 'Raxmatov', 1000000, 400000,  100000),
#         (3, 'Valeev', 150000, 60000, 15000)
#         ],
# [
#         ('',),
#         ('№', 'FIO', 'Summa', '40%',  '10%'),
#         (1, 'Qwertrg fefse', 100000, 40000,  10000),
#         (2, 'Some person', 10000, 4000,  1000),
#         (3, 'This is test', 50000, 20000, 5000),
#         (4, 'Another user', 123400, 49360, 12340)
#         ]
#     ]
#
#
# for i in data:
#     aa = sheet.cell(sheet.max_row-1, 1)
#     aa.value = '300.199'
#     aa.style = bold_style
#     sheet.row_dimensions[sheet.max_row-1].font = styles.Font(bold=True)
#     for j in i:
#         sheet.append(j)
#     sheet.append(['',
#                   'Итого',
#                   f'= SUM(C{sheet.max_row + 3 - len(i)}:C{sheet.max_row}) ',
#                   f'= SUM(D{sheet.max_row + 3 - len(i)}:D{sheet.max_row}) ',
#                   f'= SUM(G{sheet.max_row + 3 - len(i)}:G{sheet.max_row}) '
#                   ])
#     # sheet[f'D{sheet.max_row}'] =
#     # sheet[f'G{sheet.max_row}'] =
#     aa = sheet.row_dimensions[sheet.max_row]
#     aa.font = styles.Font(bold=True)
#     aa.alignment = styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
#     sheet.merge_cells(f'A{sheet.max_row + 2}:D{sheet.max_row + 3}')
#
#
# style_dict = {
#     "A": 5,
#     "B": 30,
#     "C": 13,
#     "D": 13,
#     "G": 13
# }
#
# for x, y in style_dict.items():
#     sheet.column_dimensions[x].width = y
#     sheet.column_dimensions[x].alignment = styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
#
# # sheet['D3'].alignment = styles.Alignment(horizontal="center", vertical="center")
# sheet.insert_cols(5)
# sheet.insert_cols(5)
# workbook.save(path)


thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))


class ExcelSheet:
    def __init__(self, filename: str) -> None:
        """
        :param filename: Excel file with this name will be created
        :param query_set: Django QuerySet object to fill cells
        """
        self.filename = filename
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active

    def header_style(self, row_, size=10, bold=True, is_temp=True):
        """
        :param is_temp:
        :param row_:
        :param size: Size of text in cell default size=11
        :param bold: If this param is True text  will be bold default value is True
        :return: Bold style for cell
        """
        if is_temp:
            for i in range(1, 6):
                aa = self.sheet.cell(row_, i)
                aa.font = styles.Font(bold=bold, size=size)
                aa.alignment = styles.Alignment(horizontal='center', vertical='center')
                aa.border = thin_border
        else:
            for i in range(1, 5):
                aa = self.sheet.cell(row_, i)
                aa.font = styles.Font(bold=bold, size=size)
                aa.alignment = styles.Alignment(horizontal='center', vertical='center')
                aa.border = thin_border

    def summ(self, data_: list):
        if len(data_) == 1:
            return [f"C{data_[0]}", f"D{data_[0]}", f"E{data_[0]}"]
        elif len(data_) == 2:
            return [f"{i}{data_[0]}+{i}{data_[1]}" for i in ["C", "D", "G"]]
        elif len(data_) > 3:
            ss = [f"C{data_[0]}", f"D{data_[0]}", f"G{data_[0]}"]
            for i in data_[1:]:
                ss[0] = f"{ss[0]}+C{i}"
                ss[1] = f"{ss[1]}+D{i}"
                ss[2] = f"{ss[2]}+G{i}"
            return ss

    def center_content(self):
        column_dict = {"A": 5, "B": 25, "C": 13, "D": 13, "E": 5, "F": 5, "G": 13}
        for column, width in column_dict.items():
            self.sheet.column_dimensions[column].width = width
            self.sheet.column_dimensions[column].alignment = styles.Alignment(horizontal="center", vertical="center")

    def fill_the_sheet(self, query_set):
        self.sheet.merge_cells(f'A{self.sheet.max_row}:D{self.sheet.max_row + 1}')
        temp = []
        for key, value in query_set.items():
            aa = self.sheet.cell(self.sheet.max_row - 1, 1)
            aa.value = f"{key[:3]}.{key[3:]}"
            self.header_style(self.sheet.max_row - 1, is_temp=False)
            self.header_style(self.sheet.max_row, is_temp=False)
            self.sheet.append(('',))
            self.sheet.append(('№', 'ФИО', 'Сумма', '40%', '10%'))
            self.header_style(self.sheet.max_row)
            counter = 0
            c = 0
            for j in value:
                for i in j["services"]:
                    if i['fee'] == 0:
                        continue
                    counter += 1
                    name = f"{j['secondname'] if i['fee'] == 40 else j['secondname']+' '+i['category']}"
                    pr = i["price"]
                    c += 1
                    self.sheet.append([c, name, pr, pr*i['fee']*10**-2, pr*0.1 if i["fee"] == 40 else pr*0.05])
                    self.header_style(self.sheet.max_row, bold=False)

            self.sheet.append(['',
                              'Итого',
                               f'= SUM(C{self.sheet.max_row  - counter}:C{self.sheet.max_row}) ',
                               f'= SUM(D{self.sheet.max_row  - counter}:D{self.sheet.max_row}) ',
                               f'= SUM(G{self.sheet.max_row  - counter}:G{self.sheet.max_row}) '
                               ])
            temp.append(self.sheet.max_row)
            self.header_style(self.sheet.max_row)
            self.sheet.merge_cells(f'A{self.sheet.max_row + 2}:D{self.sheet.max_row + 3}')
        self.sheet.unmerge_cells(f'A{self.sheet.max_row -1 }:D{self.sheet.max_row }')
        temp = self.summ(temp)
        self.sheet.append(['', 'Итого', f'= {temp[0]} ', f'= {temp[1]} ', f'= {temp[2]} '])
        self.header_style(self.sheet.max_row)

    def save(self):
        self.center_content()
        self.sheet.insert_cols(5)
        self.sheet.insert_cols(5)
        self.workbook.save(filename=self.filename)

