import datetime
import os
import time

import openpyxl
from openpyxl import styles
from openpyxl.drawing.image import Image
from django.conf import settings


class XlsxStyles:

    def right_style(self):
        right = styles.NamedStyle('right_positiontext')
        right.font = styles.Font(size=70, name="Tahoma")
        right.alignment = styles.Alignment(horizontal='right')
        return right

    def left_style(self):
        left = styles.NamedStyle(name="tmlefttext")
        left.font = styles.Font(size=70, name="Tahoma")
        left.alignment = styles.Alignment(horizontal='left')
        return left

    def center_style(self):
        result_style = styles.NamedStyle(name='result_text')
        result_style.font = styles.Font(size=70, name="Tahoma")
        result_style.alignment = styles.Alignment(horizontal='center', vertical='center')
        return result_style

    def center_bold_style(self):
        result_style = styles.NamedStyle('result_bold_text')
        result_style.font = styles.Font(size=70, name="Tahoma", bold=True)
        result_style.alignment = styles.Alignment(horizontal='center', vertical='center')
        return result_style


class GenerateXlsx(XlsxStyles):
    def __init__(self, input_path: str):
        self.left = self.left_style()
        self.right = self.right_style()
        self.center = self.center_style()
        self.center_bold = self.center_bold_style()
        self.number_format = "### ### ##0.00"
        self.wb = openpyxl.load_workbook(input_path)
        self.ws = self.wb.active
        self.qrcode = Image(os.path.join(settings.BASE_DIR, 'utils/images/1.png'))

    def main(self, rec_name, data):
        o_path = os.path.join(settings.BASE_DIR, 'mediafiles/kassa/{}.xlsx'.format(str(time.time())))
        start, r = 12, 12
        res = 14
        bb = self.ws.cell(6, 2)
        bb.style = self.right
        bb.value = datetime.datetime.now().strftime("%d.%m.%Y")
        # aa = self.ws.cell(7, 2)
        # aa.style = self.right
        # aa.value = rec_name
        for i in data:
            aa = self.ws.cell(r, 1)
            aa.style = self.left
            if len(i[0]) > 13:
                aa.value = f"{i[0][:14]}..."
            else:
                aa.value = i[0]
            bb = self.ws.cell(r, 2)
            bb.style = self.center
            bb.number_format = self.number_format
            bb.value = i[1]
            r += 1
            res += 1
            self.ws.insert_rows(r, 1)

        ss = self.ws.cell(res, 2)
        ss.style = self.center_bold
        ss.number_format = self.number_format
        ss.value = f"= SUM(B{start}:B{r}) "

        self.qrcode.width = 1760
        self.qrcode.height = 1200
        temp = self.ws.max_row+2
        self.ws.add_image(self.qrcode, anchor=f"A{temp}")

        self.wb.save(o_path)
        self.wb.close()
        return o_path.replace(str(settings.BASE_DIR), "")

    def check_for_patient(self, i_path, rec_name, pat_name, data):
        o_path = os.path.join(settings.BASE_DIR, 'mediafiles/kassa/{}.xlsx'.format(str(time.time())))
        workbook = openpyxl.load_workbook(i_path)
        sheet = workbook.active
        start = 8
        bb = sheet.cell(4, 2)
        bb.style = self.right
        bb.value = datetime.datetime.now().strftime("%d.%m.%Y")
        # aa = sheet.cell(5, 2)
        # aa.style = self.right
        # aa.value = rec_name
        cc = sheet.cell(5, 2)
        cc.style = self.right
        cc.value = pat_name
        for i in data:
            aa = sheet.merge_cells(f"A{start}:B{start}")
            aa = sheet.cell(start, 1)
            aa.style = self.left
            if len(i[0]) > 35:
                aa.value = f"{i[0][:34]}..."
            else:
                aa.value = i[0]
            start += 1
        workbook.save(o_path)
        workbook.close()
        return o_path.replace(str(settings.BASE_DIR), "")


i_path = os.path.join(settings.BASE_DIR, 'utils/file_templates/example.xlsx')
i_path2 = os.path.join(settings.BASE_DIR, 'utils/file_templates/patient.xlsx')
